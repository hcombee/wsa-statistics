#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
#  wsa_track_stats.py
#
#  Copyright 2019 Hans Combee (hanscombee@gmail.com)
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#
#
import time
import os
import gzip
import csv
from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from glob import glob
from datetime import datetime

front_page = [['Client time','Total time that the client was waiting until his request was fulfilled'],
['Hit time','Time that the WSA is using to fetch content from the local cache'],
['Miss time','Time that the WSA takes to fetch all Data from the server'],
['Server Transaction time','Time for the total transaction to the server to be finished'],
['Server wait time','Time until WSA gets the first byte from the Server'],
['DNS Time','Time for the WSA to do a DNS Resolution'],
['Auth Helper Wait','Time to wait for an authentication request until its validated from the AD/LDAP'],
['Auth Helper Service','Time until an authentication request is fully validated'],
['WBRS Service Time','Time for the WSA to check the reputation score'],
['Webcat Service time','Time for the WSA to check the URL Category'],
['AVC Header Scan Service Time','Time to check the Header of a request against the AVC Signatures'],
['AVC Body Scan Service time','Time to check the body of a request against the AVC Signatures'],
['Sophos/McAfee/Webroot Service Time','Time that the Scanner used to scan the object'],
['Service Queue Time','Time that the object stayed in the queue to be scanned'],
['Adaptive Scanning Service Time','Time for the adaptive scanning process to scan an object']]

header_l = ['Date/Time',1.0,1.6,2.5,4.0,6.3,10,15.8,25.1,39.8,63.1,100,158.5,251.2,398.1,631,1000,1584.9,2511.9,3981.1,6309.6,'Total requests']
header_h = ['Date/Time',10,14.6,21.4,31.3,45.7,66.9,97.8,143,209.1,305.8,447.2,654,956.4,1398.5,2045.1,2990.7,4373.4,6395.5,9352.5,13676.6,'Total requests']
wb = Workbook()

def process_file(proc_file):
	""" reads the logs line by line and puts each counter in a seperate file"""
	print("Processing ",proc_file)

	with gzip.open(proc_file,'rt') as f:
		count=1
		line = f.readline()
		while line:
			if 'Current Date' in line:
				timestamp = datetime.strptime(line[19:39],"%d %b %Y %H:%M:%S")
				client_list = [timestamp]
				hit_list = [timestamp]
				miss_list = [timestamp]
				stt_list = [timestamp]
				swt_list = [timestamp]
				dns_list = [timestamp]
				auth_list = [timestamp]
				auth_serv_list = [timestamp]
				wbrs_wait_list = [timestamp]
				wbrs_serv_list = [timestamp]
				webroot_rqh_list = [timestamp]
				webroot_rsb_list = [timestamp]
				mcafee_res_list = [timestamp]
				sophos_res_list = [timestamp]
								
			elif 'Client Time' in line:
				q_count = int(line.split()[4])
				client_list.append(q_count)
				if len(client_list) == 21:
					client_reqs = 0
					for x in range(1, len(client_list)):
						client_reqs = client_reqs + client_list[x]
					client_list.append(client_reqs)
					write_stats('client_time_stats.csv',client_list)
			elif 'Hit Time' in line:
				q_count = int(line.split()[4])
				hit_list.append(q_count)
				if len(hit_list) == 21:
					hit_reqs = 0
					for x in range(1, len(hit_list)):
						hit_reqs = hit_reqs + hit_list[x]
					hit_list.append(hit_reqs)
					write_stats('hit_time_stats.csv',hit_list)
			elif 'Miss Time' in line:
				q_count = int(line.split()[4])
				miss_list.append(q_count)
				if len(miss_list) == 21:
					miss_reqs = 0
					for x in range(1, len(miss_list)):
						miss_reqs = miss_reqs + miss_list[x]
					miss_list.append(miss_reqs)
					write_stats('miss_time_stats.csv',miss_list)
			elif 'Server Transaction Time' in line:
				q_count = int(line.split()[5])
				stt_list.append(q_count)
				if len(stt_list) == 21:
					stt_reqs = 0
					for x in range(1, len(stt_list)):
						stt_reqs = stt_reqs + stt_list[x]
					stt_list.append(stt_reqs)
					write_stats('stt_time_stats.csv',stt_list)
			elif 'Server Wait Time' in line:
				q_count = int(line.split()[5])
				swt_list.append(q_count)
				if len(swt_list) == 21:
					swt_reqs = 0
					for x in range(1, len(swt_list)):
						swt_reqs = swt_reqs + swt_list[x]
					swt_list.append(swt_reqs)
					write_stats('swt_time_stats.csv',swt_list)
			elif 'DNS Time' in line:
				q_count = int(line.split()[4])
				dns_list.append(q_count)
				if len(dns_list) == 21:
					dns_reqs = 0
					for x in range(1, len(dns_list)):
						dns_reqs = dns_reqs + dns_list[x]
					dns_list.append(dns_reqs)
					write_stats('dns_stats.csv',dns_list)
			elif 'Auth Helper Wait Time' in line:
				q_count = int(line.split()[6])
				auth_list.append(q_count)
				if len(auth_list) == 21:
					auth_reqs = 0
					for x in range(1, len(auth_list)):
						auth_reqs = auth_reqs + auth_list[x]
					auth_list.append(auth_reqs)
					write_stats('auth_helper_wait.csv',auth_list)
			elif 'Auth Helper Service Time' in line:
				q_count = int(line.split()[6])
				auth_serv_list.append(q_count)
				if len(auth_serv_list) == 21:
					auth_serv_reqs = 0
					for x in range(1, len(auth_serv_list)):
						auth_serv_reqs = auth_serv_reqs + auth_serv_list[x]
					auth_serv_list.append(auth_serv_reqs)
					write_stats('auth_serv_helper_wait.csv',auth_serv_list)
			elif 'WBRS Wait Time' in line:
				q_count = int(line.split()[5])
				wbrs_wait_list.append(q_count)
				if len(wbrs_wait_list) == 21:
					wbrs_wait_reqs = 0
					for x in range(1, len(wbrs_wait_list)):
						wbrs_wait_reqs = wbrs_wait_reqs + wbrs_wait_list[x]
					wbrs_wait_list.append(wbrs_wait_reqs)
					write_stats('wbrs_wait_stats.csv',wbrs_wait_list)
			elif 'WBRS Service Time' in line:
				q_count = int(line.split()[5])
				wbrs_serv_list.append(q_count)
				if len(wbrs_serv_list) == 21:
					wbrs_serv_reqs = 0
					for x in range(1, len(wbrs_serv_list)):
						wbrs_serv_reqs = wbrs_serv_reqs + wbrs_serv_list[x]
					wbrs_serv_list.append(wbrs_serv_reqs)
					write_stats('wbrs_serv_stats.csv',wbrs_serv_list)
			elif 'Webroot Request Header Service Time' in line:
				q_count = int(line.split()[7])
				webroot_rqh_list.append(q_count)
				if len(webroot_rqh_list) == 21:
					webroot_rqh_reqs = 0
					for x in range(1, len(webroot_rqh_list)):
						webroot_rqh_reqs = webroot_rqh_reqs + webroot_rqh_list[x]
					webroot_rqh_list.append(webroot_rqh_reqs)
					write_stats('webroot_rqh_stats.csv',webroot_rqh_list)
			elif 'Webroot Response Body Service Time' in line:
				q_count = int(line.split()[7])
				webroot_rsb_list.append(q_count)
				if len(webroot_rsb_list) == 21:
					webroot_rsb_reqs = 0
					for x in range(1, len(webroot_rsb_list)):
						webroot_rsb_reqs = webroot_rsb_reqs + webroot_rsb_list[x]
					webroot_rsb_list.append(webroot_rsb_reqs)
					write_stats('webroot_rsb_stats.csv',webroot_rsb_list)
			elif 'McAfee Response Body Service Time' in line:
				q_count = int(line.split()[7])
				mcafee_res_list.append(q_count)
				if len(mcafee_res_list) == 21:
					mcafee_res_reqs = 0
					for x in range(1, len(mcafee_res_list)):
						mcafee_res_reqs = mcafee_res_reqs + mcafee_res_list[x]
					mcafee_res_list.append(mcafee_res_reqs)
					write_stats('mcafee_res_stats.csv',mcafee_res_list)
			elif 'Sophos Response Body Service Time' in line:
				q_count = int(line.split()[7])
				sophos_res_list.append(q_count)
				if len(sophos_res_list) == 21:
					sophos_res_reqs = 0
					for x in range(1, len(sophos_res_list)):
						sophos_res_reqs = sophos_res_reqs + sophos_res_list[x]
					sophos_res_list.append(sophos_res_reqs)
					write_stats('sophos_res_stats.csv',sophos_res_list)
			
			line = f.readline()
			count += 1

def write_stats(filename,line):
	""" writes the 5 minute polls to a csv file"""
	with open(filename,'a') as f:
		writer = csv.writer(f)
		writer.writerow(line)
#		print(line)

def deltas(filename,worksheet,header):
	""" reads the csv file and calculates the deltas between two consecutive values"""
	print ("Filling",worksheet,"worksheet")
	with open(filename,'r') as f:
		read_csv = csv.reader(f)
		first = []
		for row in read_csv:
			second = row
			if not first:
				""" skip calculation for first row"""
				first = second
				write_xlsx(worksheet,header)
			elif int(second[1]) < int(first[1]):
				""" in case of a log file rollover (second row counters lower than first) """
				delta = []
				delta.append(second[0])
				for x in range (1, len(second)):
					delta.append(0)
				write_xlsx(worksheet,delta)
				first = second
			else:
				""" subtrack value of the first row from the second row 
				after wich the second row becomes the first row for the next
				calcultation"""
				delta = []
				delta.append(second[0])
				for x in range (1, len(second)):
					delta.append(int(second[x]) - int(first[x]))
				write_xlsx(worksheet,delta)
				first = second

def write_xlsx(worksheet,row):
                """ puts the calculated deltas in an Excel Workbook with conditional formatting """
		ws = wb[worksheet]
		rule = DataBarRule(start_type='percentile', start_value=10, end_type='percentile', end_value='90',color="FF638EC6", showValue="None", minLength=None, maxLength=None)
		ws.append(row)
		mr = ws.max_row
		cell_range = "B"+str(mr)+":U"+str(mr)
		ws.conditional_formatting.add(cell_range, rule)        
        

def main(args):
	print("")
	print("Reading Cisco WSA prox_track.log files")
	print("")
	file_list = glob(os.path.join('.', 'prox_track.log*'))
	output_exists = os.path.isfile('./dns_stats.csv')
	file_num=20
#	print(file_list)
	if not file_list:
		print("No Cisco WSA \"prox_track.log.x.gz\" files found")
		print("Run this script in a directory containing these files")
		print("")
	elif output_exists:
		print("Found existing statistics, using them instead")
		time.sleep(2)
	else:
		print("Processing log files, depending on hardware this can take a few minutes")
		print("")
#		print(file_list)
		while file_num >= 0:
			if file_num == 0:
				proc_file = "./prox_track.log.gz"
			else:
				proc_file = "./prox_track.log."+str(file_num)+".gz"
			if proc_file in file_list:
				process_file(proc_file)
			file_num = file_num - 1
	print(" ")
	print("Creating Worksheets")
	print(" ")
	ws = wb.active
	ws.title = "README"
	for line in front_page:       
		ws.append(line)
	ws1 = wb.create_sheet("Client Time")
	ws2 = wb.create_sheet("Hit Time")
	ws3 = wb.create_sheet("Miss Time")
	ws4 = wb.create_sheet("Server Transaction")
	ws5 = wb.create_sheet("Server Wait")
	ws6 = wb.create_sheet("DNS")
	ws7 = wb.create_sheet("Auth Helper Wait")
	ws8 = wb.create_sheet("Auth Helper Service")
	ws9 = wb.create_sheet("WBRS Wait")
	ws10 = wb.create_sheet("WBRS Service")
	ws11 = wb.create_sheet("Webroot Request Header")
	ws12 = wb.create_sheet("Webroot Response Body")
	ws13 = wb.create_sheet("McAfee Response")
	ws14 = wb.create_sheet("Sophos Response")
	deltas('client_time_stats.csv','Client Time',header_l)
	deltas('hit_time_stats.csv','Hit Time',header_l)
	deltas('miss_time_stats.csv','Miss Time',header_l)
	deltas('stt_time_stats.csv','Server Transaction',header_l)
	deltas('swt_time_stats.csv','Server Wait',header_l)
	deltas('dns_stats.csv','DNS',header_l)
	deltas('auth_helper_wait.csv','Auth Helper Wait',header_l)
	deltas('auth_serv_helper_wait.csv','Auth Helper Service',header_l)
	deltas('wbrs_wait_stats.csv','WBRS Wait',header_l)
	deltas('wbrs_serv_stats.csv','WBRS Service',header_l)
	deltas('webroot_rqh_stats.csv','Webroot Request Header',header_h)
	deltas('webroot_rsb_stats.csv','Webroot Response Body',header_h)
	deltas('mcafee_res_stats.csv','McAfee Response',header_h)
	deltas('sophos_res_stats.csv','Sophos Response',header_h)
	print('Saving "wsa_statistics.xlsx"')
	wb.save('wsa_statistics.xlsx')
	return 0

if __name__ == '__main__':
    import sys
    sys.exit(main(sys.argv))
